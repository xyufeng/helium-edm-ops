from __future__ import annotations

import argparse
import csv
import html
import json
import os
import re
import sys
import textwrap
import time
from dataclasses import asdict, dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
TAG_RE = re.compile(r"<[^>]+>")
WS_RE = re.compile(r"\s+")
HREF_RE = re.compile(r"(?is)<a\b[^>]*\bhref=[\"']([^\"']+)[\"'][^>]*>")
IMG_RE = re.compile(r"(?is)<img\b([^>]*)>")
PLACEHOLDER_VALUES = {
    "replace_me",
    "change_me",
    "replace_with_random_secret",
    "your_sendy_api_key",
    "your_emaillistverify_api_key",
}


@dataclass(frozen=True)
class Contact:
    email: str
    name: str = ""
    source_row: int = 0


@dataclass(frozen=True)
class VerifiedContact:
    email: str
    name: str
    status: str
    disposition: str
    accepted: bool
    reason: str


@dataclass(frozen=True)
class SuppressionEntry:
    email: str
    reason: str = "suppressed"


@dataclass(frozen=True)
class FileAssessment:
    path: str
    role: str
    confidence: float
    reason: str


@dataclass(frozen=True)
class IntakePlan:
    client: str
    contacts_path: str
    edm_html_path: str
    header_path: str
    footer_path: str
    subject: str
    actions: list[str]


@dataclass(frozen=True)
class ClientConfig:
    slug: str
    display_name: str = ""
    sendy_brand_id: str = ""
    sendy_brand_name: str = ""
    sendy_list_id: str = ""
    from_name: str = ""
    from_email: str = ""
    reply_to: str = ""
    header_path: str = ""
    footer_path: str = ""
    required_footer_text: str = ""
    dashboard_visible: bool = True


def env(name: str, default: str = "") -> str:
    value = os.environ.get(name, default).strip()
    return "" if value.lower() in PLACEHOLDER_VALUES else value


def load_dotenv(path: Path = Path(".env")) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        key, value = stripped.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def normalize_email(value: str) -> str:
    return value.strip().lower()


def read_contacts(csv_path: Path, email_column: str = "email", name_column: str = "name") -> tuple[list[Contact], list[str]]:
    warnings: list[str] = []
    contacts: list[Contact] = []

    with csv_path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError("CSV has no header row.")

        fields = {field.strip().lower(): field for field in reader.fieldnames}
        email_field = fields.get(email_column.lower())
        name_field = fields.get(name_column.lower())

        if not email_field:
            possible = ", ".join(reader.fieldnames)
            raise ValueError(f"Email column '{email_column}' not found. Available columns: {possible}")

        seen: set[str] = set()
        for row_number, row in enumerate(reader, start=2):
            email_address = normalize_email(row.get(email_field, ""))
            name = (row.get(name_field, "") if name_field else "").strip()

            if not email_address:
                warnings.append(f"row {row_number}: blank email skipped")
                continue
            if not EMAIL_RE.match(email_address):
                warnings.append(f"row {row_number}: malformed email skipped: {email_address}")
                continue
            if email_address in seen:
                warnings.append(f"row {row_number}: duplicate skipped: {email_address}")
                continue

            seen.add(email_address)
            contacts.append(Contact(email=email_address, name=name, source_row=row_number))

    return contacts, warnings


def read_suppression_list(path: Path | None) -> dict[str, SuppressionEntry]:
    if not path or not path.exists():
        return {}
    entries: dict[str, SuppressionEntry] = {}
    if path.suffix.lower() == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        rows = data if isinstance(data, list) else data.get("suppressions", []) if isinstance(data, dict) else []
        for row in rows:
            if isinstance(row, str):
                email_address = normalize_email(row)
                reason = "suppressed"
            elif isinstance(row, dict):
                email_address = normalize_email(str(row.get("email", "")))
                reason = str(row.get("reason", "suppressed")).strip() or "suppressed"
            else:
                continue
            if email_address:
                entries[email_address] = SuppressionEntry(email=email_address, reason=reason)
        return entries

    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return {}
        fields = {field.strip().lower(): field for field in reader.fieldnames}
        email_field = fields.get("email") or fields.get("email address") or fields.get("e-mail")
        reason_field = fields.get("reason")
        if not email_field:
            return {}
        for row in reader:
            email_address = normalize_email(row.get(email_field, ""))
            if not email_address:
                continue
            reason = (row.get(reason_field, "") if reason_field else "").strip() or "suppressed"
            entries[email_address] = SuppressionEntry(email=email_address, reason=reason)
    return entries


def apply_suppression(verified: list[VerifiedContact], suppressions: dict[str, SuppressionEntry]) -> list[VerifiedContact]:
    if not suppressions:
        return verified
    updated: list[VerifiedContact] = []
    for contact in verified:
        suppression = suppressions.get(contact.email)
        if suppression and contact.accepted:
            updated.append(
                VerifiedContact(
                    email=contact.email,
                    name=contact.name,
                    status=contact.status,
                    disposition="suppressed",
                    accepted=False,
                    reason=f"suppressed: {suppression.reason}",
                )
            )
        else:
            updated.append(contact)
    return updated


def html_to_plain_text(html_text: str) -> str:
    body = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", html_text)
    body = re.sub(r"(?i)<br\s*/?>", "\n", body)
    body = re.sub(r"(?i)</p\s*>", "\n\n", body)
    body = TAG_RE.sub(" ", body)
    body = html.unescape(body)
    lines = [WS_RE.sub(" ", line).strip() for line in body.splitlines()]
    return "\n".join(line for line in lines if line).strip()


def body_inner_html(html_text: str) -> str:
    match = re.search(r"(?is)<body[^>]*>(.*?)</body>", html_text)
    if match:
        return match.group(1).strip()
    return html_text.strip()


def title_from_html(html_text: str) -> str:
    for pattern in (r"(?is)<title[^>]*>(.*?)</title>", r"(?is)<h1[^>]*>(.*?)</h1>"):
        match = re.search(pattern, html_text)
        if match:
            value = html_to_plain_text(match.group(1))
            if value:
                return value
    return ""


def render_helium_email(edm_html: str, header_html: str, footer_html: str) -> str:
    body = "\n".join(part for part in [body_inner_html(header_html), body_inner_html(edm_html), body_inner_html(footer_html)] if part)
    return (
        '<!doctype html>\n'
        '<html>\n'
        '  <head>\n'
        '    <meta charset="utf-8">\n'
        '    <meta name="viewport" content="width=device-width, initial-scale=1">\n'
        '  </head>\n'
        '  <body style="margin:0; padding:0; background:#f6f7f9;">\n'
        '    <div style="max-width:680px; margin:0 auto; background:#ffffff; font-family:Arial, sans-serif; color:#111827;">\n'
        f'{body}\n'
        '    </div>\n'
        '  </body>\n'
        '</html>\n'
    )


def slugify_client(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.strip().lower()).strip("-")
    return slug or "default"


def resolve_client_templates(client: str, header: Path | None, footer: Path | None) -> tuple[Path, Path]:
    client_slug = slugify_client(client)
    config = load_client_config(client_slug)
    client_dir = Path("templates") / "clients" / client_slug
    client_header = client_dir / "header.html"
    client_footer = client_dir / "footer.html"

    config_header = Path(config.header_path) if config.header_path else None
    config_footer = Path(config.footer_path) if config.footer_path else None

    header_path = header or config_header or client_header
    footer_path = footer or config_footer or client_footer

    if not header and not client_header.exists():
        header_path = Path("templates/clients/default/header.html")
    if not footer and not client_footer.exists():
        footer_path = Path("templates/clients/default/footer.html")

    return header_path, footer_path


def load_client_config(client: str) -> ClientConfig:
    client_slug = slugify_client(client)
    path = Path("config") / "clients" / f"{client_slug}.json"
    if not path.exists():
        return ClientConfig(slug=client_slug)
    data = json.loads(path.read_text(encoding="utf-8"))
    return ClientConfig(
        slug=client_slug,
        display_name=str(data.get("display_name", client_slug)).strip(),
        sendy_brand_id=str(data.get("sendy_brand_id", "")).strip(),
        sendy_brand_name=str(data.get("sendy_brand_name", "")).strip(),
        sendy_list_id=str(data.get("sendy_list_id", "")).strip(),
        from_name=str(data.get("from_name", "")).strip(),
        from_email=str(data.get("from_email", "")).strip(),
        reply_to=str(data.get("reply_to", "")).strip(),
        header_path=str(data.get("header_path", "")).strip(),
        footer_path=str(data.get("footer_path", "")).strip(),
        required_footer_text=str(data.get("required_footer_text", "")).strip(),
        dashboard_visible=bool(data.get("dashboard_visible", True)),
    )


def parse_sendy_list_ids(value: str) -> list[str]:
    return [item.strip() for item in value.split(",") if item.strip()]


def parse_decimal(value: str | int | float | Decimal) -> Decimal:
    try:
        return Decimal(str(value or "0")).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return Decimal("0.00")


def parse_rate(value: str | int | float | Decimal) -> Decimal:
    try:
        return Decimal(str(value or "0"))
    except (InvalidOperation, ValueError):
        return Decimal("0")


def format_money(value: Decimal, currency: str) -> str:
    return f"{currency} {value:.2f}"


def format_rate(value: Decimal) -> str:
    formatted = format(value.normalize(), "f")
    return formatted if "." not in formatted else formatted.rstrip("0").rstrip(".")


def write_invoice_artifacts(
    output_dir: Path,
    summary: dict[str, Any],
    *,
    partner: str = "",
    currency: str = "SGD",
    campaign_fee: str = "0",
    verification_unit_fee: str = "0",
    list_fee: str = "0",
    sending_unit_fee: str = "0",
    commission_rate: str = "0.5",
    discount: str = "0",
    period: str = "",
) -> dict[str, Any]:
    currency = (currency or "SGD").strip().upper()
    partner = partner.strip() or summary.get("client", "Partner")
    period = period.strip() or time.strftime("%b %Y")
    campaign_amount = parse_decimal(campaign_fee)
    verification_rate = parse_rate(verification_unit_fee)
    sending_rate = parse_rate(sending_unit_fee or list_fee)
    commission_decimal = parse_rate(commission_rate)
    discount_amount = parse_decimal(discount)
    contacts_read = int(summary.get("contacts_read", 0) or 0)
    list_count = len(summary.get("sendy_list_ids") or [])
    setup_units = 1 if summary.get("sendy_campaign_mode") != "disabled" else 0
    sending_units = int(summary.get("sendy_imported_contacts", 0) or 0) * max(list_count, 1)

    line_items = [
        {
            "description": "Setup Cost",
            "quantity": setup_units,
            "unit_price": campaign_amount,
            "amount": (campaign_amount * setup_units).quantize(Decimal("0.01")),
        },
        {
            "description": "Email Cleaning",
            "quantity": contacts_read,
            "unit_price": verification_rate,
            "amount": (verification_rate * contacts_read).quantize(Decimal("0.01")),
        },
        {
            "description": "Email Sending",
            "quantity": sending_units,
            "unit_price": sending_rate,
            "amount": (sending_rate * sending_units).quantize(Decimal("0.01")),
        },
    ]
    total = (sum((item["amount"] for item in line_items), Decimal("0.00")) - discount_amount).quantize(Decimal("0.01"))
    commission = (total * commission_decimal).quantize(Decimal("0.01"))
    payable = (total - commission).quantize(Decimal("0.01"))
    invoice_id = f"HE-{time.strftime('%Y%m%d-%H%M%S')}"
    invoice = {
        "invoice_id": invoice_id,
        "date": time.strftime("%Y-%m-%d"),
        "period": period,
        "partner": partner,
        "currency": currency,
        "client": summary.get("client", ""),
        "subject": summary.get("subject", ""),
        "mode": summary.get("mode", ""),
        "contacts_read": contacts_read,
        "sending_units": sending_units,
        "setup_units": setup_units,
        "accepted": summary.get("accepted", 0),
        "rejected": summary.get("rejected", 0),
        "quarantined": summary.get("quarantined", 0),
        "suppressed": summary.get("suppressed", 0),
        "sendy_lists": summary.get("sendy_list_id", ""),
        "discount": str(discount_amount),
        "commission_rate": str(commission_decimal),
        "line_items": [
            {
                **item,
                "unit_price": str(item["unit_price"]),
                "amount": str(item["amount"]),
            }
            for item in line_items
        ],
        "total": str(total),
        "commission": str(commission),
        "payable": str(payable),
    }

    (output_dir / "invoice.json").write_text(json.dumps(invoice, indent=2), encoding="utf-8")
    with (output_dir / "invoice_rows.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Helium Emails", "", invoice_id])
        writer.writerow(["Invoice ID", invoice_id])
        writer.writerow(["Date", invoice["date"]])
        writer.writerow(["Period", period])
        writer.writerow(["Partner", partner])
        writer.writerow(["Client", invoice["client"]])
        writer.writerow(["Campaign", invoice["subject"]])
        writer.writerow(["Mode", invoice["mode"]])
        writer.writerow(["Blast Setup Cost per campaign", "", format_money(campaign_amount, currency)])
        writer.writerow(["Sending Cost per email", "", format_rate(sending_rate)])
        writer.writerow(["Cleaning Cost per email", "", format_rate(verification_rate)])
        writer.writerow([])
        writer.writerow(["Description", "Quantity", "Unit price", "Amount"])
        for item in line_items:
            unit_price = format_money(item["unit_price"], currency) if item["description"] == "Setup Cost" else format_rate(item["unit_price"])
            writer.writerow([item["description"], item["quantity"], unit_price, format_money(item["amount"], currency)])
        writer.writerow([])
        writer.writerow(["DISCOUNT", "", "", format_money(discount_amount, currency)])
        writer.writerow(["Total Cost", "", "", format_money(total, currency)])
        writer.writerow(["Commission", "", "", format_money(commission, currency)])
        writer.writerow(["PAYABLE", "", "", format_money(payable, currency)])
        writer.writerow([f"ALL PRICES ARE IN {currency}"])

    rows = "\n".join(
        "<tr>"
        f"<td>{html.escape(item['description'])}</td>"
        f"<td>{item['quantity']}</td>"
        f"<td>{html.escape(format_money(item['unit_price'], currency) if item['description'] == 'Setup Cost' else format_rate(item['unit_price']))}</td>"
        f"<td>{html.escape(format_money(item['amount'], currency))}</td>"
        "</tr>"
        for item in line_items
    )
    invoice_html = textwrap.dedent(
        f"""
        <!doctype html>
        <html>
          <head>
            <meta charset="utf-8">
            <title>{html.escape(invoice_id)} Invoice</title>
            <style>
              body {{ font-family: Arial, sans-serif; color: #111827; margin: 40px; }}
              h1 {{ margin-bottom: 4px; }}
              table {{ border-collapse: collapse; width: 100%; margin-top: 24px; }}
              th, td {{ border-bottom: 1px solid #d1d5db; padding: 10px; text-align: left; }}
              th {{ background: #f3f4f6; }}
              .total {{ font-size: 22px; font-weight: 700; text-align: right; margin-top: 24px; }}
              .meta {{ color: #4b5563; line-height: 1.6; }}
            </style>
          </head>
          <body>
            <h1>Invoice {html.escape(invoice_id)}</h1>
            <div class="meta">
              <div>Date: {html.escape(invoice["date"])}</div>
              <div>Period: {html.escape(period)}</div>
              <div>Partner: {html.escape(partner)}</div>
              <div>Client: {html.escape(str(invoice["client"]))}</div>
              <div>Campaign: {html.escape(str(invoice["subject"]))}</div>
              <div>Mode: {html.escape(str(invoice["mode"]))}</div>
            </div>
            <table>
              <thead><tr><th>Description</th><th>Quantity</th><th>Unit price</th><th>Amount</th></tr></thead>
              <tbody>{rows}</tbody>
            </table>
            <div class="total">Total Cost: {html.escape(format_money(total, currency))}</div>
            <div class="total">Commission: {html.escape(format_money(commission, currency))}</div>
            <div class="total">Payable: {html.escape(format_money(payable, currency))}</div>
            <p class="meta">ALL PRICES ARE IN {html.escape(currency)}</p>
          </body>
        </html>
        """
    ).strip()
    (output_dir / "invoice.html").write_text(invoice_html, encoding="utf-8")
    return invoice


def csv_has_email_column(path: Path) -> tuple[bool, str]:
    try:
        with path.open(newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader, [])
            normalized = [item.strip().lower() for item in header]
            if "email" in normalized or "email address" in normalized or "e-mail" in normalized:
                return True, "CSV header contains an email column."
            sample = "\n".join(",".join(row) for _, row in zip(range(8), reader))
            if re.search(r"[^@\s]+@[^@\s]+\.[^@\s]+", sample):
                return True, "CSV sample contains email addresses."
    except UnicodeDecodeError:
        return False, "Could not read as UTF-8 CSV."
    except OSError as exc:
        return False, f"Could not inspect CSV: {exc}"
    return False, "CSV does not look like a contact list."


def safe_sheet_slug(value: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "-", value.strip()).strip("-").lower()
    return slug or "sheet"


def worksheet_has_data(rows: list[list[Any]]) -> bool:
    return any(any(cell not in (None, "") for cell in row) for row in rows)


def convert_xlsx_to_csvs(path: Path, output_dir: Path | None = None) -> list[Path]:
    output_dir = output_dir or path.parent
    workbook = load_workbook(path, read_only=True, data_only=True)
    converted: list[Path] = []
    used_names: set[str] = set()
    try:
        for worksheet in workbook.worksheets:
            rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
            if not worksheet_has_data(rows):
                continue
            base_name = f"{path.stem}__{safe_sheet_slug(worksheet.title)}"
            csv_name = base_name
            suffix = 2
            while csv_name in used_names:
                csv_name = f"{base_name}-{suffix}"
                suffix += 1
            used_names.add(csv_name)
            csv_path = output_dir / f"{csv_name}.csv"
            with csv_path.open("w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerows(rows)
            converted.append(csv_path)
    finally:
        workbook.close()
    return converted


def convert_xlsx_files_in_dir(input_dir: Path) -> list[Path]:
    converted: list[Path] = []
    for path in sorted(input_dir.iterdir()):
        if path.is_file() and path.suffix.lower() == ".xlsx":
            converted.extend(convert_xlsx_to_csvs(path, input_dir))
    return converted


def classify_file(path: Path) -> FileAssessment:
    name = path.name.lower()
    suffix = path.suffix.lower()
    if suffix == ".csv":
        has_email, reason = csv_has_email_column(path)
        if has_email:
            return FileAssessment(str(path), "contacts", 0.95, reason)
        return FileAssessment(str(path), "unknown", 0.2, reason)

    if suffix in {".html", ".htm"}:
        try:
            content = path.read_text(encoding="utf-8", errors="replace")
        except OSError as exc:
            return FileAssessment(str(path), "unknown", 0.0, f"Could not read file: {exc}")

        if "header" in name:
            return FileAssessment(str(path), "header", 0.95, "Filename indicates this is the header.")
        if "footer" in name:
            return FileAssessment(str(path), "footer", 0.95, "Filename indicates this is the footer.")
        if any(token in content.lower() for token in ("<html", "<body", "<table", "<p", "<h1")):
            return FileAssessment(str(path), "edm_html", 0.8, "HTML content looks like the campaign body.")
        return FileAssessment(str(path), "unknown", 0.3, "HTML file did not look like an EDM body.")

    if suffix in {".json", ".txt", ".md"}:
        return FileAssessment(str(path), "notes", 0.65, "Text-like file that may contain campaign context.")

    return FileAssessment(str(path), "unknown", 0.1, f"Unsupported file extension: {suffix or '(none)'}")


def extract_subject_from_notes(path: Path) -> str:
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return ""

    if path.suffix.lower() == ".json":
        try:
            data = json.loads(text)
        except json.JSONDecodeError:
            data = {}
        for key in ("subject", "subject_line", "campaign_subject", "title"):
            value = data.get(key) if isinstance(data, dict) else ""
            if isinstance(value, str) and value.strip():
                return value.strip()

    for line in text.splitlines():
        match = re.match(r"(?i)\s*(subject|subject line|title)\s*:\s*(.+?)\s*$", line)
        if match:
            return match.group(2).strip()
    return ""


def choose_assessment(assessments: list[FileAssessment], role: str) -> FileAssessment | None:
    matches = [item for item in assessments if item.role == role]
    if not matches:
        return None
    return sorted(matches, key=lambda item: item.confidence, reverse=True)[0]


def build_intake_plan(input_dir: Path, subject: str, client: str, default_header: Path, default_footer: Path) -> tuple[IntakePlan, list[FileAssessment]]:
    if not input_dir.exists():
        raise ValueError(f"Input directory does not exist: {input_dir}")

    converted_csvs = convert_xlsx_files_in_dir(input_dir)
    files = [path for path in sorted(input_dir.iterdir()) if path.is_file()]
    assessments = [classify_file(path) for path in files]

    contacts = choose_assessment(assessments, "contacts")
    edm = choose_assessment(assessments, "edm_html")
    header = choose_assessment(assessments, "header")
    footer = choose_assessment(assessments, "footer")
    notes = [Path(item.path) for item in assessments if item.role == "notes"]

    if not contacts:
        raise ValueError("No contact CSV found. Add a CSV with an email column to the intake folder.")
    if not edm:
        raise ValueError("No EDM HTML found. Add the client EDM HTML to the intake folder.")

    inferred_subject = subject
    if not inferred_subject:
        for note in notes:
            inferred_subject = extract_subject_from_notes(note)
            if inferred_subject:
                break
    if not inferred_subject:
        inferred_subject = title_from_html(Path(edm.path).read_text(encoding="utf-8", errors="replace"))
    if not inferred_subject:
        inferred_subject = f"Helium EDM campaign {time.strftime('%Y-%m-%d')}"

    header_path = Path(header.path) if header else default_header
    footer_path = Path(footer.path) if footer else default_footer
    actions = [
        "Classify uploaded files by role.",
        *([f"Convert {len(converted_csvs)} worksheet(s) from uploaded XLSX files into CSV files."] if converted_csvs else []),
        "Use contact CSV as the recipient list.",
        "Use EDM HTML as the campaign body.",
        f"Apply the {client} client header and footer unless uploaded files override them.",
        "Generate plain-text fallback and run AI preflight.",
        "Verify contacts through EmailListVerify.",
        "Upload accepted contacts to Sendy.",
        "Create a Sendy draft campaign ready to be reviewed and sent.",
    ]

    return (
        IntakePlan(
            client=client,
            contacts_path=str(Path(contacts.path)),
            edm_html_path=str(Path(edm.path)),
            header_path=str(header_path),
            footer_path=str(footer_path),
            subject=inferred_subject,
            actions=actions,
        ),
        assessments,
    )


def verify_email(api_key: str, email_address: str, timeout: int = 30) -> str:
    response = requests.get(
        "https://apps.emaillistverify.com/api/verifyEmail",
        params={"secret": api_key, "email": email_address},
        timeout=timeout,
    )
    response.raise_for_status()
    return response.text.strip().lower()


DEFAULT_ACCEPTED_STATUSES = {"ok"}
DEFAULT_QUARANTINE_STATUSES = {"unknown", "risky", "catch_all", "catch-all", "accept_all", "accept-all"}


def verify_contacts(
    contacts: list[Contact],
    accepted_statuses: set[str],
    quarantine_statuses: set[str],
    dry_run: bool,
    pause_seconds: float,
) -> list[VerifiedContact]:
    api_key = env("EMAILLISTVERIFY_API_KEY")
    if not dry_run and not api_key:
        raise ValueError("EMAILLISTVERIFY_API_KEY is required unless --dry-run is set.")

    verified: list[VerifiedContact] = []
    for contact in contacts:
        if dry_run:
            status = "ok" if not contact.email.endswith(".invalid") else "invalid"
        else:
            status = verify_email(api_key, contact.email)
            if pause_seconds:
                time.sleep(pause_seconds)

        accepted = status in accepted_statuses
        disposition = "accepted" if accepted else "quarantine" if status in quarantine_statuses else "rejected"
        if disposition == "accepted":
            reason = "accepted"
        elif disposition == "quarantine":
            reason = f"quarantined for human review by EmailListVerify status '{status}'"
        else:
            reason = f"rejected by EmailListVerify status '{status}'"
        verified.append(
            VerifiedContact(
                email=contact.email,
                name=contact.name,
                status=status,
                disposition=disposition,
                accepted=accepted,
                reason=reason,
            )
        )
    return verified


class SendyClient:
    def __init__(self, base_url: str, api_key: str, dry_run: bool) -> None:
        self.base_url = base_url.rstrip("/")
        self.api_key = api_key
        self.dry_run = dry_run

    def post(self, path: str, payload: dict[str, Any]) -> str:
        if self.dry_run:
            return f"DRY_RUN {path} {json.dumps(redact(payload), sort_keys=True)}"
        response = requests.post(f"{self.base_url}{path}", data=payload, timeout=60)
        response.raise_for_status()
        return response.text.strip()

    def subscribe(self, list_id: str, contact: VerifiedContact) -> str:
        return self.post(
            "/subscribe",
            {
                "api_key": self.api_key,
                "list": list_id,
                "email": contact.email,
                "name": contact.name,
                "boolean": "true",
            },
        )

    def get_brands(self) -> Any:
        result = self.post("/api/brands/get-brands.php", {"api_key": self.api_key})
        return parse_sendy_json(result)

    def get_lists(self, brand_id: str, include_hidden: bool = False) -> Any:
        result = self.post(
            "/api/lists/get-lists.php",
            {
                "api_key": self.api_key,
                "brand_id": brand_id,
                "include_hidden": "yes" if include_hidden else "no",
            },
        )
        return parse_sendy_json(result)

    def create_campaign(
        self,
        *,
        from_name: str,
        from_email: str,
        reply_to: str,
        title: str,
        subject: str,
        html_text: str,
        plain_text: str,
        list_ids: str,
        brand_id: str,
        send_campaign: bool,
        schedule_date_time: str,
        schedule_timezone: str,
    ) -> str:
        payload = {
            "api_key": self.api_key,
            "from_name": from_name,
            "from_email": from_email,
            "reply_to": reply_to,
            "title": title,
            "subject": subject,
            "html_text": html_text,
            "plain_text": plain_text,
            "track_opens": "1",
            "track_clicks": "1",
            "send_campaign": "1" if send_campaign else "0",
        }
        if send_campaign:
            payload["list_ids"] = list_ids
        else:
            payload["brand_id"] = brand_id
        if schedule_date_time:
            payload["schedule_date_time"] = schedule_date_time
            payload["schedule_timezone"] = schedule_timezone
        return self.post("/api/campaigns/create.php", payload)


def redact(payload: dict[str, Any]) -> dict[str, Any]:
    redacted = dict(payload)
    if "api_key" in redacted:
        redacted["api_key"] = "***"
    return redacted


def parse_sendy_json(value: str) -> Any:
    if value.lower().startswith("error:"):
        raise ValueError(value)
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        try:
            return json.loads(value, strict=False)
        except json.JSONDecodeError as exc:
            raise ValueError(f"Unexpected Sendy response: {value}") from exc


def ai_preflight(html_text: str, plain_text: str, subject: str, client_note: str) -> dict[str, Any]:
    api_key = env("OPENAI_API_KEY")
    if not api_key:
        return {
            "enabled": False,
            "note": "OPENAI_API_KEY not set; used deterministic fallback.",
            "suggested_subject": subject,
            "plain_text_preview": plain_text[:600],
            "checks": deterministic_checks(html_text, plain_text, subject),
        }

    from openai import OpenAI

    client = OpenAI(api_key=api_key)
    prompt = {
        "role": "user",
        "content": textwrap.dedent(
            f"""
            You are helping Helium.sg prepare an EDM campaign received from a client.
            Return strict JSON with:
            - suggested_subject: one concise subject line
            - risk_flags: array of deliverability/compliance/content risks
            - fixes: array of concrete edits to make before sending
            - plain_text_summary: 4-6 sentence plain-text fallback summary

            Client note:
            {client_note or "(none)"}

            Current subject:
            {subject}

            HTML:
            {html_text[:12000]}
            """
        ).strip(),
    }
    response = client.responses.create(
        model=env("OPENAI_MODEL", "gpt-4.1-mini"),
        input=[prompt],
        text={"format": {"type": "json_object"}},
    )
    return json.loads(response.output_text)


def deterministic_checks(html_text: str, plain_text: str, subject: str) -> list[str]:
    return [item["message"] for item in deliverability_checks(html_text, plain_text, subject)]


def deliverability_checks(html_text: str, plain_text: str, subject: str, required_footer_text: str = "") -> list[dict[str, str]]:
    structured: list[dict[str, str]] = []
    lower_html = html_text.lower()

    def add(severity: str, code: str, message: str) -> None:
        structured.append({"severity": severity, "code": code, "message": message})

    if "[unsubscribe]" not in lower_html and "unsubscribe" not in lower_html:
        add("error", "missing_unsubscribe", "No obvious unsubscribe text/tag found.")
    if len(subject) > 80:
        add("warning", "long_subject", "Subject is longer than 80 characters.")
    if not plain_text:
        add("error", "missing_plain_text", "Plain text version is empty.")

    for href in HREF_RE.findall(html_text):
        normalized = href.strip().lower()
        if normalized.startswith("http://"):
            add("warning", "non_https_link", f"Non-HTTPS link detected: {href.strip()}")
        if not normalized or normalized == "#":
            add("warning", "empty_link", "Empty or placeholder link detected.")

    for img_attrs in IMG_RE.findall(html_text):
        if not re.search(r"(?is)\balt\s*=", img_attrs):
            add("warning", "missing_image_alt", "Image tag is missing alt text.")

    if len(html_text.encode("utf-8")) > 100_000:
        add("warning", "large_html", "HTML is larger than 100 KB.")

    if required_footer_text and required_footer_text.lower() not in lower_html:
        add("error", "missing_required_footer_text", f"Required footer text not found: {required_footer_text}")

    return structured or [{"severity": "ok", "code": "ok", "message": "No deterministic issues found."}]


def write_report(
    output_dir: Path,
    verified: list[VerifiedContact],
    warnings: list[str],
    ai_result: dict[str, Any],
    sendy_results: list[dict[str, str]],
    campaign_result: str,
    dry_run: bool,
    import_to_sendy: bool,
    create_campaign: bool,
    sendy_list_ids: list[str] | None = None,
    sendy_brand_id: str = "",
    invoice: dict[str, Any] | None = None,
    consent_basis: str = "",
    consent_confirmed: bool = False,
    plan: IntakePlan | None = None,
    assessments: list[FileAssessment] | None = None,
    rendered_html_path: Path | None = None,
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    accepted = [item for item in verified if item.accepted]
    rejected = [item for item in verified if not item.accepted]
    quarantined = [item for item in verified if item.disposition == "quarantine"]
    suppressed = [item for item in verified if item.disposition == "suppressed"]
    rejected_only = [item for item in verified if item.disposition == "rejected"]

    with (output_dir / "verified_contacts.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["email", "name", "status", "disposition", "accepted", "reason"])
        writer.writeheader()
        for item in verified:
            writer.writerow(asdict(item))

    report = {
        "summary": {
            "mode": "dry_run" if dry_run else "live",
            "accepted": len(accepted),
            "rejected": len(rejected_only),
            "quarantined": len(quarantined),
            "suppressed": len(suppressed),
            "warnings": len(warnings),
        },
        "file_assessments": [asdict(item) for item in assessments or []],
        "processing_plan": asdict(plan) if plan else None,
        "consent": {
            "confirmed": consent_confirmed,
            "basis": consent_basis,
        },
        "external_actions": {
            "email_verification": "simulated" if dry_run else "emaillistverify_live",
            "sendy_import": "disabled" if not import_to_sendy else "simulated" if dry_run else "sendy_live",
            "sendy_campaign": "disabled" if not create_campaign else "simulated" if dry_run else "sendy_live",
            "sendy_brand_id": sendy_brand_id,
            "sendy_list_ids": sendy_list_ids or [],
        },
        "rendered_html_path": str(rendered_html_path) if rendered_html_path else None,
        "warnings": warnings,
        "ai_preflight": ai_result,
        "sendy_import": sendy_results,
        "campaign_result": campaign_result,
        "invoice": invoice or {},
    }
    (output_dir / "run_report.json").write_text(json.dumps(report, indent=2), encoding="utf-8")


def process_campaign(
    *,
    contacts_path: Path,
    html_path: Path,
    header_path: Path,
    footer_path: Path,
    subject: str,
    client: str,
    list_id: str,
    brand_id: str,
    output_dir: Path,
    from_name: str = "",
    from_email: str = "",
    reply_to: str = "",
    title: str = "",
    client_note: str = "",
    consent_basis: str = "",
    consent_confirmed: bool = False,
    suppression_path: Path | None = None,
    email_column: str = "email",
    name_column: str = "name",
    accepted_statuses: set[str] | None = None,
    quarantine_statuses: set[str] | None = None,
    dry_run: bool = False,
    import_to_sendy: bool = False,
    create_campaign: bool = False,
    send_campaign: bool = False,
    schedule_date_time: str = "",
    schedule_timezone: str = "Asia/Singapore",
    verification_pause: float = 0.0,
    invoice_partner: str = "",
    invoice_currency: str = "SGD",
    invoice_campaign_fee: str = "0",
    invoice_verification_unit_fee: str = "0",
    invoice_list_fee: str = "0",
    invoice_sending_unit_fee: str = "0",
    invoice_commission_rate: str = "0.5",
    invoice_discount: str = "0",
    invoice_period: str = "",
    plan: IntakePlan | None = None,
    assessments: list[FileAssessment] | None = None,
) -> dict[str, Any]:
    accepted_statuses = accepted_statuses or DEFAULT_ACCEPTED_STATUSES
    quarantine_statuses = quarantine_statuses or DEFAULT_QUARANTINE_STATUSES
    consent_basis = consent_basis.strip() or ("provided_client_consent" if consent_confirmed else "")
    client_config = load_client_config(client)
    list_id = list_id or client_config.sendy_list_id
    list_ids = parse_sendy_list_ids(list_id)
    list_id = ",".join(list_ids)
    brand_id = brand_id or client_config.sendy_brand_id
    from_name = from_name or client_config.from_name or env("SENDY_DEFAULT_FROM_NAME", "Helium")
    from_email = from_email or client_config.from_email or env("SENDY_DEFAULT_FROM_EMAIL", "hello@helium.sg")
    reply_to = reply_to or client_config.reply_to or env("SENDY_DEFAULT_REPLY_TO", "hello@helium.sg")

    if not dry_run and (import_to_sendy or create_campaign) and not consent_confirmed:
        raise ValueError("Confirm that the uploaded list has provided consent before live Sendy actions.")
    if import_to_sendy and not list_ids:
        raise ValueError("At least one Sendy list is required for contact upload.")
    if create_campaign and not brand_id and not send_campaign:
        raise ValueError("Sendy brand ID is required to create a draft campaign.")

    edm_html = html_path.read_text(encoding="utf-8")
    header_html = header_path.read_text(encoding="utf-8") if header_path.exists() else ""
    footer_html = footer_path.read_text(encoding="utf-8") if footer_path.exists() else ""
    html_text = render_helium_email(edm_html, header_html, footer_html)
    output_dir.mkdir(parents=True, exist_ok=True)
    rendered_html_path = output_dir / "rendered_edm.html"
    rendered_html_path.write_text(html_text, encoding="utf-8")

    plain_text = html_to_plain_text(html_text)
    contacts, warnings = read_contacts(contacts_path, email_column, name_column)
    verified = verify_contacts(contacts, accepted_statuses, quarantine_statuses, dry_run, verification_pause)
    suppressions = read_suppression_list(suppression_path)
    verified = apply_suppression(verified, suppressions)
    accepted = [item for item in verified if item.accepted]
    quarantined = [item for item in verified if item.disposition == "quarantine"]
    suppressed = [item for item in verified if item.disposition == "suppressed"]
    rejected_only = [item for item in verified if item.disposition == "rejected"]

    ai_result = ai_preflight(html_text, plain_text, subject, client_note)
    delivery_checks = deliverability_checks(html_text, plain_text, subject, client_config.required_footer_text)
    ai_result["deterministic_checks"] = delivery_checks
    blocking_checks = [item for item in delivery_checks if item["severity"] == "error"]
    if not dry_run and create_campaign and blocking_checks:
        messages = "; ".join(item["message"] for item in blocking_checks)
        raise ValueError(f"Blocking deliverability checks failed: {messages}")

    sendy = SendyClient(env("SENDY_BASE_URL"), env("SENDY_API_KEY"), dry_run)
    if not dry_run and (import_to_sendy or create_campaign):
        if not env("SENDY_BASE_URL") or not env("SENDY_API_KEY"):
            raise ValueError("SENDY_BASE_URL and SENDY_API_KEY are required for Sendy calls.")

    sendy_results: list[dict[str, str]] = []
    if import_to_sendy:
        for selected_list_id in list_ids:
            for contact in accepted:
                result = sendy.subscribe(selected_list_id, contact)
                sendy_results.append({"email": contact.email, "list_id": selected_list_id, "result": result})

    campaign_result = "skipped"
    if create_campaign:
        if not send_campaign and not brand_id:
            raise ValueError("--brand-id is required when creating a Sendy draft.")
        campaign_result = sendy.create_campaign(
            from_name=from_name,
            from_email=from_email,
            reply_to=reply_to,
            title=title or subject,
            subject=ai_result.get("suggested_subject", subject),
            html_text=html_text,
            plain_text=ai_result.get("plain_text_summary", plain_text),
            list_ids=list_id,
            brand_id=brand_id,
            send_campaign=send_campaign,
            schedule_date_time=schedule_date_time,
            schedule_timezone=schedule_timezone,
        )

    summary = {
        "subject": subject,
        "client": client,
        "contacts_file": str(contacts_path),
        "edm_file": str(html_path),
        "contacts_read": len(contacts),
        "accepted": len(accepted),
        "rejected": len(rejected_only),
        "quarantined": len(quarantined),
        "suppressed": len(suppressed),
        "warnings": len(warnings),
        "rendered_html": str(rendered_html_path),
        "mode": "dry_run" if dry_run else "live",
        "email_verification": "simulated" if dry_run else "emaillistverify_live",
        "sendy_import_mode": "disabled" if not import_to_sendy else "simulated" if dry_run else "sendy_live",
        "sendy_campaign_mode": "disabled" if not create_campaign else "simulated" if dry_run else "sendy_live",
        "campaign_result": campaign_result,
        "sendy_list_id": list_id,
        "sendy_list_ids": list_ids,
        "sendy_brand_id": brand_id,
        "from_name": from_name,
        "from_email": from_email,
        "reply_to": reply_to,
        "consent_basis": consent_basis,
        "consent_confirmed": consent_confirmed,
        "suppression_file": str(suppression_path) if suppression_path else "",
        "output_dir": str(output_dir),
        "sendy_imported": len(sendy_results),
        "sendy_imported_contacts": len(accepted) if sendy_results else 0,
    }
    invoice = write_invoice_artifacts(
        output_dir,
        summary,
        partner=invoice_partner,
        currency=invoice_currency,
        campaign_fee=invoice_campaign_fee,
        verification_unit_fee=invoice_verification_unit_fee,
        list_fee=invoice_list_fee,
        sending_unit_fee=invoice_sending_unit_fee,
        commission_rate=invoice_commission_rate,
        discount=invoice_discount,
        period=invoice_period,
    )
    summary["invoice_id"] = invoice["invoice_id"]
    summary["invoice_total"] = invoice["total"]
    summary["invoice_commission"] = invoice["commission"]
    summary["invoice_payable"] = invoice["payable"]
    summary["invoice_currency"] = invoice["currency"]
    summary["invoice_partner"] = invoice["partner"]
    summary["invoice_period"] = invoice["period"]

    write_report(
        output_dir,
        verified,
        warnings,
        ai_result,
        sendy_results,
        campaign_result,
        dry_run,
        import_to_sendy,
        create_campaign,
        sendy_list_ids=list_ids,
        sendy_brand_id=brand_id,
        invoice=invoice,
        consent_basis=consent_basis,
        consent_confirmed=consent_confirmed,
        plan=plan,
        assessments=assessments,
        rendered_html_path=rendered_html_path,
    )

    return summary


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Automate Helium.sg list verification and Sendy campaign setup.",
    )
    parser.add_argument("--input-dir", type=Path, help="Folder of uploaded client files for the agent to classify and process.")
    parser.add_argument("--client", default="default", help="Helium client slug used for client-specific header/footer templates.")
    parser.add_argument("--contacts", type=Path, help="CSV list from client.")
    parser.add_argument("--html", type=Path, help="EDM HTML file from client.")
    parser.add_argument("--header", type=Path, help="Header HTML to prepend. Overrides the client template.")
    parser.add_argument("--footer", type=Path, help="Footer HTML to append. Overrides the client template.")
    parser.add_argument("--subject", default="", help="Campaign subject line.")
    parser.add_argument("--title", default="", help="Internal Sendy campaign title.")
    parser.add_argument("--client-note", default="", help="Context for AI preflight.")
    parser.add_argument("--consent-basis", default="", help="Audit note for why this list is permissioned.")
    parser.add_argument("--confirm-consent", action="store_true", help="Confirm uploaded list has provided consent. Required for live Sendy actions.")
    parser.add_argument("--suppression-list", type=Path, help="CSV or JSON suppression list. Suppressed accepted contacts are not uploaded.")
    parser.add_argument("--list-id", default="", help="Sendy list ID(s) to import/send to. Use a comma-separated value for multiple lists. Defaults from client config when available.")
    parser.add_argument("--brand-id", default="", help="Sendy brand ID for draft creation.")
    parser.add_argument("--from-name", default="", help="Override sender name.")
    parser.add_argument("--from-email", default="", help="Override sender email.")
    parser.add_argument("--reply-to", default="", help="Override reply-to email.")
    parser.add_argument("--email-column", default="email")
    parser.add_argument("--name-column", default="name")
    parser.add_argument("--accepted-status", action="append", default=sorted(DEFAULT_ACCEPTED_STATUSES), help="Accepted EmailListVerify status. Repeatable.")
    parser.add_argument("--quarantine-status", action="append", default=sorted(DEFAULT_QUARANTINE_STATUSES), help="EmailListVerify status to quarantine for review. Repeatable.")
    parser.add_argument("--output-dir", type=Path, default=Path("runs/latest"))
    parser.add_argument("--dry-run", action="store_true", help="Do not call external APIs.")
    parser.add_argument("--import-to-sendy", action="store_true", help="Subscribe accepted contacts to Sendy.")
    parser.add_argument("--create-campaign", action="store_true", help="Create campaign in Sendy.")
    parser.add_argument("--send-campaign", action="store_true", help="Send immediately instead of creating a draft.")
    parser.add_argument("--schedule-date-time", default="", help='Example: "June 15, 2026 6:05pm".')
    parser.add_argument("--schedule-timezone", default="Asia/Singapore")
    parser.add_argument("--verification-pause", type=float, default=0.0, help="Seconds between EmailListVerify calls.")
    parser.add_argument("--invoice-partner", default="", help="Partner/customer name for invoice artifacts.")
    parser.add_argument("--invoice-currency", default="SGD", help="Invoice currency code.")
    parser.add_argument("--invoice-campaign-fee", default="0", help="Flat campaign operations fee.")
    parser.add_argument("--invoice-verification-unit-fee", default="0", help="Per-contact verification fee.")
    parser.add_argument("--invoice-list-fee", default="0", help="Per selected Sendy list upload fee.")
    parser.add_argument("--invoice-sending-unit-fee", default="0", help="Per-email sending fee. Defaults separately from verification.")
    parser.add_argument("--invoice-commission-rate", default="0.5", help="Commission share as a decimal. Default: 0.5.")
    parser.add_argument("--invoice-discount", default="0", help="Invoice discount amount.")
    parser.add_argument("--invoice-period", default="", help='Invoice period, for example "Dec 2024 - Feb 2025".')
    return parser


def ensure_input_paths(args: argparse.Namespace) -> tuple[Path, Path, Path, Path, str, IntakePlan | None, list[FileAssessment]]:
    client = slugify_client(args.client)
    default_header, default_footer = resolve_client_templates(client, args.header, args.footer)

    if args.input_dir:
        plan, assessments = build_intake_plan(args.input_dir, args.subject, client, default_header, default_footer)
        return (
            Path(plan.contacts_path),
            Path(plan.edm_html_path),
            Path(plan.header_path),
            Path(plan.footer_path),
            plan.subject,
            plan,
            assessments,
        )

    missing = []
    if not args.contacts:
        missing.append("--contacts")
    if not args.html:
        missing.append("--html")
    if not args.subject:
        missing.append("--subject")
    if missing:
        joined = ", ".join(missing)
        raise ValueError(f"Missing required arguments without --input-dir: {joined}")

    plan = IntakePlan(
        client=client,
        contacts_path=str(args.contacts),
        edm_html_path=str(args.html),
        header_path=str(default_header),
        footer_path=str(default_footer),
        subject=args.subject,
        actions=[
            "Use explicitly provided contacts and HTML files.",
            f"Apply the {client} client header and footer.",
            "Generate plain-text fallback and run AI preflight.",
            "Verify contacts through EmailListVerify.",
            "Upload accepted contacts to Sendy if requested.",
            "Create a Sendy campaign if requested.",
        ],
    )
    assessments = [
        FileAssessment(str(args.contacts), "contacts", 1.0, "Provided explicitly with --contacts."),
        FileAssessment(str(args.html), "edm_html", 1.0, "Provided explicitly with --html."),
        FileAssessment(str(default_header), "header", 1.0, "Resolved from explicit override or client template."),
        FileAssessment(str(default_footer), "footer", 1.0, "Resolved from explicit override or client template."),
    ]
    return args.contacts, args.html, default_header, default_footer, args.subject, plan, assessments


def main(argv: list[str] | None = None) -> int:
    load_dotenv()
    args = build_parser().parse_args(argv)

    contacts_path, html_path, header_path, footer_path, subject, plan, assessments = ensure_input_paths(args)

    summary = process_campaign(
        contacts_path=contacts_path,
        html_path=html_path,
        header_path=header_path,
        footer_path=footer_path,
        subject=subject,
        client=plan.client if plan else slugify_client(args.client),
        list_id=args.list_id,
        brand_id=args.brand_id,
        output_dir=args.output_dir,
        from_name=args.from_name,
        from_email=args.from_email,
        reply_to=args.reply_to,
        title=args.title,
        client_note=args.client_note,
        consent_basis=args.consent_basis,
        consent_confirmed=args.confirm_consent,
        suppression_path=args.suppression_list,
        email_column=args.email_column,
        name_column=args.name_column,
        accepted_statuses=set(args.accepted_status),
        quarantine_statuses=set(args.quarantine_status),
        dry_run=args.dry_run,
        import_to_sendy=args.import_to_sendy,
        create_campaign=args.create_campaign,
        send_campaign=args.send_campaign,
        schedule_date_time=args.schedule_date_time,
        schedule_timezone=args.schedule_timezone,
        verification_pause=args.verification_pause,
        invoice_partner=args.invoice_partner,
        invoice_currency=args.invoice_currency,
        invoice_campaign_fee=args.invoice_campaign_fee,
        invoice_verification_unit_fee=args.invoice_verification_unit_fee,
        invoice_list_fee=args.invoice_list_fee,
        invoice_sending_unit_fee=args.invoice_sending_unit_fee,
        invoice_commission_rate=args.invoice_commission_rate,
        invoice_discount=args.invoice_discount,
        invoice_period=args.invoice_period,
        plan=plan,
        assessments=assessments,
    )
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
