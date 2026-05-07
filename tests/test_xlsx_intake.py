from pathlib import Path

from openpyxl import Workbook

from helium_edm.cli import build_intake_plan, convert_xlsx_to_csvs, read_contacts, render_helium_email


def write_workbook(path: Path) -> None:
    workbook = Workbook()
    contacts = workbook.active
    contacts.title = "Consented Leads"
    contacts.append(["name", "email", "company"])
    contacts.append(["Alice Tan", "alice@example.com", "Example Co"])
    contacts.append(["Ben Lim", "ben@example.com", "Helium"])

    notes = workbook.create_sheet("Notes")
    notes.append(["Campaign", "May briefing"])
    notes.append(["Owner", "Helium"])

    empty = workbook.create_sheet("Empty Sheet")
    empty.append([None, None])

    workbook.save(path)


def test_convert_xlsx_to_csvs_writes_one_csv_per_non_empty_sheet(tmp_path):
    workbook_path = tmp_path / "client-list.xlsx"
    write_workbook(workbook_path)

    converted = convert_xlsx_to_csvs(workbook_path)

    assert [path.name for path in converted] == [
        "client-list__consented-leads.csv",
        "client-list__notes.csv",
    ]
    assert "alice@example.com" in converted[0].read_text(encoding="utf-8")


def test_convert_xlsx_to_csvs_is_idempotent(tmp_path):
    workbook_path = tmp_path / "client-list.xlsx"
    write_workbook(workbook_path)

    first = convert_xlsx_to_csvs(workbook_path)
    second = convert_xlsx_to_csvs(workbook_path)

    assert [path.name for path in first] == [path.name for path in second]
    assert not list(tmp_path.glob("*-2.csv"))


def test_build_intake_plan_classifies_generated_xlsx_sheet_csv(tmp_path):
    workbook_path = tmp_path / "client-list.xlsx"
    html_path = tmp_path / "campaign.html"
    header_path = tmp_path / "header.html"
    footer_path = tmp_path / "footer.html"
    write_workbook(workbook_path)
    html_path.write_text("<html><body><h1>May briefing</h1><p>Hello</p></body></html>", encoding="utf-8")
    header_path.write_text("<header>Header</header>", encoding="utf-8")
    footer_path.write_text("<footer>Footer</footer>", encoding="utf-8")

    plan, assessments = build_intake_plan(tmp_path, "", "test", header_path, footer_path)

    assert plan.contacts_path.endswith("client-list__consented-leads.csv")
    assert Path(plan.contacts_path).exists()
    assert any(item.path.endswith("client-list.xlsx") and item.role == "unknown" for item in assessments)
    assert any(item.path.endswith("client-list__notes.csv") and item.role == "unknown" for item in assessments)


def test_read_contacts_accepts_chinese_email_header(tmp_path):
    csv_path = tmp_path / "contacts.csv"
    csv_path.write_text("姓名,电子邮件,公司\n王小明,wang@example.com,Acme\n", encoding="utf-8")

    contacts, warnings = read_contacts(csv_path)

    assert contacts[0].email == "wang@example.com"
    assert contacts[0].name == "王小明"
    assert any("电子邮件" in warning for warning in warnings)


def test_read_contacts_infers_unexpected_email_column_from_values(tmp_path):
    csv_path = tmp_path / "contacts.csv"
    csv_path.write_text("姓名,主要联系方式,公司\n王小明,wang@example.com,Acme\n", encoding="utf-8")

    contacts, warnings = read_contacts(csv_path)

    assert contacts[0].email == "wang@example.com"
    assert contacts[0].name == "王小明"
    assert any("Inferred email column" in warning for warning in warnings)


def test_render_helium_email_uses_isle_sendy_tags_and_subject(tmp_path):
    header = Path("templates/clients/isle/header.html").read_text(encoding="utf-8")
    footer = Path("templates/clients/isle/footer.html").read_text(encoding="utf-8")

    rendered = render_helium_email(
        "<html><body><p>Campaign body</p></body></html>",
        header,
        footer,
        subject="Gallop into A Fortunate Year",
    )

    assert "Gallop into A Fortunate Year" in rendered
    assert "<webversion>here to view the online version</webversion>" in rendered
    assert "[Email]" in rendered
    assert "<unsubscribe>unsubscribe here</unsubscribe>" in rendered


def test_render_helium_email_does_not_duplicate_existing_sendy_wrapper():
    edm = """
    <html><body>
      <p>Problems viewing this email? click <webversion>here</webversion>.</p>
      <p>Campaign body</p>
      <p>Please click <unsubscribe>unsubscribe here</unsubscribe>.</p>
    </body></html>
    """

    rendered = render_helium_email(edm, "<p>{{ subject }}</p><webversion>view</webversion>", "<unsubscribe>bye</unsubscribe>", subject="Subject")

    assert rendered.count("<webversion") == 1
    assert rendered.count("<unsubscribe") == 1
    assert "Subject" not in rendered
